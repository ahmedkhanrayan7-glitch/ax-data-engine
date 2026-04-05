"""
AX Engine — Export Utilities

Converts LeadResult lists to CSV and XLSX formats.
"""
from __future__ import annotations

import csv
import io
from typing import Any, Dict, List


def _flatten_lead(lead: Dict[str, Any]) -> Dict[str, Any]:
    """Flatten nested lead dict to tabular format."""
    flat = {
        "company_name": lead.get("company_name", ""),
        "location": lead.get("location", ""),
        "website": lead.get("website", ""),
        "phone": lead.get("phone", ""),
        "address": lead.get("address", ""),
        "lead_score": lead.get("lead_score", 0),
    }

    # Decision makers (top 3)
    dms = lead.get("decision_makers", [])
    for i, dm in enumerate(dms[:3]):
        flat[f"dm_{i+1}_name"] = dm.get("name", "")
        flat[f"dm_{i+1}_role"] = dm.get("role", "")
        flat[f"dm_{i+1}_confidence"] = dm.get("confidence_score", 0)

    # Contacts
    contacts = lead.get("contacts", {})
    flat["emails"] = "; ".join(contacts.get("emails", []))
    flat["phones"] = "; ".join(contacts.get("phones", []))
    flat["primary_email"] = contacts.get("primary_email", "")
    flat["email_status"] = contacts.get("email_status", "")
    flat["socials"] = "; ".join(contacts.get("socials", []))

    # Enrichment
    enrichment = lead.get("enrichment", {})
    flat["company_size"] = enrichment.get("company_size", "")
    flat["revenue_estimate"] = enrichment.get("revenue_estimate", "")
    flat["tech_stack"] = ", ".join(enrichment.get("tech_stack", []))
    flat["google_rating"] = enrichment.get("google_rating", "")
    flat["year_founded"] = enrichment.get("year_founded", "")

    # Opportunities
    signals = lead.get("opportunity_signals", [])
    flat["opportunity_signals"] = " | ".join(signals[:3])

    return flat


def results_to_csv(results: List[Dict]) -> bytes:
    """Convert leads to CSV bytes."""
    if not results:
        return b""

    flattened = [_flatten_lead(r) for r in results]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=flattened[0].keys())
    writer.writeheader()
    writer.writerows(flattened)

    return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def results_to_xlsx(results: List[Dict]) -> bytes:
    """Convert leads to XLSX bytes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl required for XLSX export: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AX Leads"

    if not results:
        return b""

    flattened = [_flatten_lead(r) for r in results]
    headers = list(flattened[0].keys())

    # Header row styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, lead in enumerate(flattened, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=lead.get(key, ""))

    # Auto-size columns
    for col in ws.columns:
        max_width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_width + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
